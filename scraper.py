import re
import time
import random
import pandas as pd
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def init_driver():
    options = uc.ChromeOptions()
    options.add_argument('--disable-logging')
    options.add_argument('--ignore-certificate-errors')
    return uc.Chrome(options=options)

def parse_salary_avg(salary_text: str) -> float:
    units = {'L': 1, 'Cr': 100}
    matches = re.findall(r'₹?([\d.]+)\s*([A-Za-z]+)', salary_text)
    if len(matches) == 2:
        low = float(matches[0][0]) * units.get(matches[0][1], 1)
        high = float(matches[1][0]) * units.get(matches[1][1], 1)
        return low + (high - low) / 2
    elif len(matches) == 1:
        return float(matches[0][0]) * units.get(matches[0][1], 1)
    return 0.0

def extract_salaries(driver, url):
    time.sleep(random.uniform(5, 8))
    driver.get(url)
    time.sleep(random.uniform(5, 8))
    rows = driver.find_elements(By.CSS_SELECTOR, 'tr[data-jobprofileid]')
    seen = set()
    results = []
    for row in rows:
        try:
            role_el = row.find_element(By.CSS_SELECTOR, 'td.left-content a p.card-content__company')
            salary_el = row.find_element(By.CSS_SELECTOR, 'p.salary-range')
            role = role_el.text.strip()
            salary = salary_el.text.strip()
            if role not in seen:
                seen.add(role)
                results.append((role, salary))
        except:
            continue
    return results

def search_and_scrape(company_name):
    depts = [
        '/engineering-software-qa-department?experience=0',
        '/data-science-analytics-department?experience=0',
        '/fresher'
    ]
    driver = init_driver()
    try:
        driver.get("https://www.google.com")
        wait = WebDriverWait(driver, 3)
        box = wait.until(EC.element_to_be_clickable((By.NAME, "q")))
        box.clear()
        box.send_keys(f"{company_name} site:ambitionbox.com salaries")
        box.send_keys(Keys.RETURN)

        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'h3')))
        driver.find_element(By.CSS_SELECTOR, 'h3').click()

        base = driver.current_url.split('-salaries/')[0]
        if '-salaries' not in base:
            base += '-salaries'

        for dept in depts:
            full_url = base + dept
            try:
                salaries = extract_salaries(driver, full_url)
                if salaries:
                    return salaries
            except TimeoutException:
                continue
        return []
    except Exception as e:
        print(f"Scraping failed for {company_name}: {e}")
        return []
    finally:
        try:
            driver.quit()
        except:
            pass


def update_excel_with_salaries(csv_file):
    try:
        df = pd.read_csv(csv_file, encoding='utf-8').dropna(subset=["Company"])
    except UnicodeDecodeError:
        try:
            df = pd.read_csv(csv_file, encoding='latin1').dropna(subset=["Company"])
        except:
            df = pd.read_csv(csv_file, encoding='cp1252').dropna(subset=["Company"])

    wb = Workbook()
    sh = wb.active
    sh.title = "Salaries"
    green = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    orange = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    blank = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    sh.append(["Company", "Top 3 Roles", "Highest Paying Role"])

    for idx, row in df.iterrows():
        company = row['Company']
        if pd.isna(company):
            print(f"Skipping row {idx + 1}: No company name.")
            continue

        print(f"Processing {company}...")
        role_data = search_and_scrape(company)

        if role_data:
            averaged_roles = [
                (r, parse_salary_avg(s)) for r, s in role_data
                if parse_salary_avg(s) > 0
            ]
            averaged_roles.sort(key=lambda x: x[1], reverse=True)
            if averaged_roles:
                top_roles = [r for r, _ in averaged_roles[:3]]
                highest_avg = averaged_roles[0][1]
                highest_role = averaged_roles[0][0]
                role_text = ", ".join(top_roles)
                sh.append([company, role_text, highest_role])
                fill = (
                    green if highest_avg >= 20 else
                    yellow if highest_avg >= 15 else
                    orange if highest_avg >= 10 else
                    red
                )
                for cell in sh[sh.max_row]:
                    cell.fill = fill
            else:
                sh.append([company, "Masked/Unavailable", "Masked/Unavailable"])
                for cell in sh[sh.max_row]:
                    cell.fill = blank
        else:
            sh.append([company, "", "No Data"])
            for cell in sh[sh.max_row]:
                cell.fill = blank

    wb.save("salaries_output.xlsx")

# Run the script directly
csv_file = r"C:\Users\aypey\Desktop\Computer_Science\ArmaanScrapingtest.csv"
update_excel_with_salaries(csv_file)
